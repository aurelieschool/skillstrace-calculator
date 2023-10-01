from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    send_file,
)
from werkzeug.utils import secure_filename
import os
import tempfile
import calc
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment

app = Flask(__name__)

file_status = {"ready": False, "data": None}


@app.route("/")
def index():
    return render_template("index.html")


def getScoresSummary(file, assessment_type):
    try:
        file_path = os.path.join(tempfile.gettempdir(), secure_filename(file.filename))
        file.save(file_path)

        file_status["data"] = calc.summaryFromFile(file_path, assessment_type)
        file_status["ready"] = True

        os.remove(file_path)

        return "File succesfully processed"
    except Exception as e:
        return f"Error processing file: {str(e)}"


def get_scores_summary(file, assessment_type):
    try:
        file_path = os.path.join(tempfile.gettempdir(), secure_filename(file.filename))
        file.save(file_path)

        file_status["data"] = calc.report_from_file(file_path, assessment_type)
        file_status["ready"] = True

        os.remove(file_path)

        return "file processed"
    except Exception as e:
        return f"Error processing file: {str(e)}"


@app.route("/processFile", methods=["POST"])
def processFile():
    file = request.files.get("file")
    assessment_type = request.form.get("assessment_type")
    if file and file.filename != "":
        message = get_scores_summary(file, assessment_type)
        if "error" in message:
            return jsonify({"error": message})
        else:
            return jsonify({"message": message})
    else:
        return jsonify({"error": "no file selected or file name empty"})


@app.route("/check_file_status", methods=["GET"])
def check_file_status():
    global file_ready
    if file_status["ready"]:
        return jsonify({"message": "done"})
    else:
        return jsonify({"message": "not done"})


@app.route("/download/<assessment_type>", methods=["GET"])
def download(assessment_type):
    if assessment_type == "ELD1_FLS" or assessment_type == "ELD2_FLS":
        header_color = "234682"
        main_color = "8bb0f0"
    elif assessment_type == "ELD1_CLS" or assessment_type == "ELD2_CLS":
        header_color = "25732d"
        main_color = "b4dbb8"
    elif assessment_type == "ELD1_LFC" or assessment_type == "ELD2_LFC":
        header_color = "ba2d90"
        main_color = "e8aed7"

    if assessment_type and file_status["ready"]:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            file_status["data"].to_excel(writer, sheet_name="Sheet1", index=False)

        # Load the workbook from the buffer
        buffer.seek(0)
        percent_style = NamedStyle(name="percent", number_format="0%")

        # Save DataFrame to Excel directly into the buffer
        # Load the workbook and select the active worksheet
        wb = load_workbook(buffer)
        ws = wb.active

        if not "percent" in wb.named_styles:
            wb.add_named_style(percent_style)
        for cell in ws["C2:C" + str(ws.max_row)]:
            cell[0].style = "percent"

        # Justify left the first column && make it all bold
        for row in ws.iter_rows(min_col=1, max_col=1):
            ws.row_dimensions[row[0].row].height = 30  # making the height big
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(bold=True)

        for row in ws.iter_rows(min_col=2, max_col=3):
            ws.row_dimensions[row[0].row].height = 30  # making the height big
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # make the font red if they did worse than 50%
        for row in ws.iter_rows():
            for cell in row:
                try:
                    value = float(cell.value)
                    if value < 0.5:
                        cell.font = Font(color="c23329")
                except (TypeError, ValueError):
                    # This will skip the cell if its content isn't a number (e.g., a string or None)
                    pass

        # Color the "Score" header cell
        score_header = [
            ws["A1"],
            ws["B1"],
            ws["C1"],
        ]  # Assuming the Score header is in cell B1
        for header in score_header:
            header.fill = PatternFill(
                start_color="BDBDBD", end_color="BDBDBD", fill_type="solid"
            )
            header.font = Font(bold=True, size=12)

        # Apply border to all columns
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # make the class score (second row) bigger:
        for cell in ws[2]:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=header_color, end_color=header_color, fill_type="solid"
            )

        # make color for the rows that are beneath row 2
        for row_index, row in enumerate(ws.iter_rows(), start=1):
            if row_index > 2:
                for cell in row:
                    cell.fill = PatternFill(
                        start_color=main_color, end_color=main_color, fill_type="solid"
                    )

        # have automatic width
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]  # Convert generator object to list
            for cell in column:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2  # Adding a little extra space
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the styled Excel file
        wb.save("styled_filename.xlsx")

        # back to the start of the buffer
        buffer.seek(0)
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{assessment_type}_summary.xlsx",
        )
    else:
        return "No file ready for download"


if __name__ == "__main__":
    app.run(debug=True)
